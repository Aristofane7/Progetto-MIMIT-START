#!/usr/bin/env python3
"""
P-TSA (OR7.4) - Codice di calcolo del Product Technological Sustainability
Assessment. Applica lo schema LCT/ISO 14040 (cradle-to-grave) alle 3 tipologie
di prodotto definite dagli EPD (7.4 / 8.2 / 20 mm) e calcola:
  indicatori  SCR (IOA) / PsI (OP) / OCR (TQ)
  sotto-indici IOAI / OPI / TQI  e  indice sintetico  P-TSI
  con due normalizzazioni:  z-score + pesi uguali (PRIMARIA)
                            scoring 1-5 + AHP (SECONDARIA, CR<=0.10)
  + TII (Technology Improvement Index) e analisi di sensibilita.

WORKFLOW "sostituisci e rilancia":
  1) i DATI DI INPUT stanno nel file  RP7.4_dataset_sintetico.xlsx  (foglio INPUT).
  2) al primo avvio, se il file non esiste, viene generato dai valori di default
     (serie provvisorie, ancorate a EPD e a RP6.x/RP7.3, in corso di consolidamento).
  3) per rifare l'assessment con i DATI REALI: sostituire i valori nel foglio
     INPUT del dataset e rilanciare  `python3 RP7.4_build.py`  -> la struttura di
     calcolo resta invariata e vengono rigenerati calculation_log, weights e figure.

Output: RP7.4_dataset_sintetico.xlsx, RP7.4_weights.xlsx,
        RP7.4_calculation_log.xlsx, RP7.4_fig1_*.png, RP7.4_fig2_*.png
"""
import os, json
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.dirname(os.path.abspath(__file__))
DATASET = os.path.join(OUT, "RP7.4_dataset_sintetico.xlsx")

TYPES = ["T1_7.4mm", "T2_8.2mm", "T3_20mm"]
PERIODS = ["t", "t-1"]                       # t = lug2023-giu2024 (EPD); t-1 = anno prec. (per TII)

# ---- colonne del dataset di INPUT (una riga per tipologia x periodo) ----
FIELDS = ["spessore_mm", "massa_kg_m2", "stabilimento",
          # IOA (SCR = stock medio / consumo medio giornaliero, in giorni)
          "stock_rawmat_t", "cons_rawmat_tgg",
          "stock_finished_m2", "cons_finished_m2gg",
          "stock_glaze_t", "cons_glaze_tgg",
          # OP (PsI)
          "firstchoice_rate", "energia_processo_MJ_m2", "resa_m2_m2", "throughput_m2_h",
          # TQ (OCR = QP / AT su parametri ISO 10545, soglie EN 14411 BIa)
          "flex_QP_Nmm2", "flex_AT_Nmm2",
          "break_QP_N", "break_AT_N",
          "surf_QP_pct", "surf_AT_pct"]

# ============================================================================
# 1. DATI DI DEFAULT (serie provvisorie, in corso di consolidamento)
#    Ancoraggio: EPD (massa/m2, parametri ISO 10545), RP7.3 (energia D060/D240).
# ============================================================================
def default_inputs():
    static = {
        "T1_7.4mm": dict(spessore_mm=7.4,  massa_kg_m2=13.98, stabilimento="D060 Scandiano"),
        "T2_8.2mm": dict(spessore_mm=8.2,  massa_kg_m2=16.05, stabilimento="D060 Scandiano"),
        "T3_20mm":  dict(spessore_mm=20.0, massa_kg_m2=41.79, stabilimento="D240 Frassinoro"),
    }
    # energia specifica di processo ~ 3.05 MJ/kg * massa (8.2mm ~ 49 MJ/m2 ~ RP7.3)
    e_spec = {t: round(3.05 * static[t]["massa_kg_m2"], 1) for t in TYPES}
    base = {  # periodo t
      "T1_7.4mm": dict(stock_rawmat_t=4200, cons_rawmat_tgg=105, stock_finished_m2=95000, cons_finished_m2gg=3200,
                       stock_glaze_t=48, cons_glaze_tgg=2.1, firstchoice_rate=0.958,
                       energia_processo_MJ_m2=e_spec["T1_7.4mm"], resa_m2_m2=0.962, throughput_m2_h=640,
                       flex_QP_Nmm2=48.0, flex_AT_Nmm2=35.0, break_QP_N=1400, break_AT_N=700,
                       surf_QP_pct=97.6, surf_AT_pct=95.0),
      "T2_8.2mm": dict(stock_rawmat_t=4600, cons_rawmat_tgg=100, stock_finished_m2=88000, cons_finished_m2gg=2600,
                       stock_glaze_t=52, cons_glaze_tgg=2.0, firstchoice_rate=0.951,
                       energia_processo_MJ_m2=e_spec["T2_8.2mm"], resa_m2_m2=0.955, throughput_m2_h=560,
                       flex_QP_Nmm2=50.0, flex_AT_Nmm2=35.0, break_QP_N=2050, break_AT_N=1300,
                       surf_QP_pct=97.1, surf_AT_pct=95.0),
      "T3_20mm":  dict(stock_rawmat_t=5200, cons_rawmat_tgg=62, stock_finished_m2=52000, cons_finished_m2gg=900,
                       stock_glaze_t=40, cons_glaze_tgg=1.2, firstchoice_rate=0.936,
                       energia_processo_MJ_m2=e_spec["T3_20mm"], resa_m2_m2=0.941, throughput_m2_h=210,
                       flex_QP_Nmm2=53.0, flex_AT_Nmm2=35.0, break_QP_N=7200, break_AT_N=1300,
                       surf_QP_pct=96.4, surf_AT_pct=95.0),
    }
    data = {}
    for t in TYPES:
        row_t = {**static[t], **base[t]}
        data[(t, "t")] = row_t
        # periodo t-1: anno precedente leggermente meno performante (per il TII)
        r1 = dict(row_t)
        r1["stock_rawmat_t"]   = round(row_t["stock_rawmat_t"]*0.97, 1)
        r1["cons_rawmat_tgg"]  = round(row_t["cons_rawmat_tgg"]*1.03, 2)
        r1["stock_finished_m2"]= round(row_t["stock_finished_m2"]*0.97, 1)
        r1["cons_finished_m2gg"]=round(row_t["cons_finished_m2gg"]*1.03, 2)
        r1["stock_glaze_t"]    = round(row_t["stock_glaze_t"]*0.97, 2)
        r1["cons_glaze_tgg"]   = round(row_t["cons_glaze_tgg"]*1.03, 3)
        r1["firstchoice_rate"] = round(row_t["firstchoice_rate"]-0.008, 3)
        r1["energia_processo_MJ_m2"] = round(row_t["energia_processo_MJ_m2"]*1.02, 2)
        r1["resa_m2_m2"]       = round(row_t["resa_m2_m2"]-0.006, 3)
        r1["throughput_m2_h"]  = round(row_t["throughput_m2_h"]*0.97, 1)
        r1["flex_QP_Nmm2"]     = round(row_t["flex_QP_Nmm2"]-0.6, 2)
        r1["surf_QP_pct"]      = round(row_t["surf_QP_pct"]-0.5, 2)
        data[(t, "t-1")] = r1
    return data

# ============================================================================
# 2. DATASET I/O
# ============================================================================
HFILL = PatternFill("solid", fgColor="4F6228"); HFONT = Font(color="FFFFFF", bold=True)
def _style_header(ws):
    for c in ws[1]:
        c.fill = HFILL; c.font = HFONT; c.alignment = Alignment(horizontal="center")
def _autosize(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(w+2, 10), 26)

def write_dataset(data, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "INPUT"
    ws.append(["tipologia", "periodo"] + FIELDS)
    for t in TYPES:
        for p in PERIODS:
            r = data[(t, p)]
            ws.append([t, p] + [r[f] for f in FIELDS])
    _style_header(ws); _autosize(ws)
    ws2 = wb.create_sheet("LEGENDA")
    legenda = [
      ("SCR (IOA)", "Stock Coverage Rate = stock medio / consumo medio giornaliero [giorni]"),
      ("PsI (OP)",  "Productivity Indicators: energia m2/GJ = firstchoice/(energia/1000); resa m2/m2; throughput m2/h"),
      ("OCR (TQ)",  "Output Conformity Rate = QP/AT (ISO 10545; soglie EN 14411 BIa)"),
      ("periodo",   "t = lug2023-giu2024 (EPD); t-1 = anno precedente (per il TII)"),
      ("NOTA",      "Serie provvisorie, in corso di consolidamento. Sostituire con i dati reali e rilanciare RP7.4_build.py."),
    ]
    for k, v in legenda: ws2.append([k, v])
    ws2.column_dimensions["A"].width = 14; ws2.column_dimensions["B"].width = 95
    wb.save(path)

def read_dataset(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb["INPUT"]
    hdr = [c.value for c in ws[1]]
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(hdr, row))
        if rec.get("tipologia") is None: continue
        key = (rec["tipologia"], rec["periodo"])
        data[key] = {f: rec[f] for f in FIELDS}
    return data

# carica il dataset se presente, altrimenti crealo dai default
if os.path.exists(DATASET):
    data = read_dataset(DATASET); print(f"[input] letto {os.path.basename(DATASET)}")
else:
    data = default_inputs(); write_dataset(data, DATASET); print(f"[input] creato {os.path.basename(DATASET)} dai default")

# ============================================================================
# 3. INDICATORI
# ============================================================================
def SCR_of(p):
    return {
      "SCR_RawMat":   {t: data[(t,p)]["stock_rawmat_t"]   / data[(t,p)]["cons_rawmat_tgg"]   for t in TYPES},
      "SCR_Finished": {t: data[(t,p)]["stock_finished_m2"]/ data[(t,p)]["cons_finished_m2gg"] for t in TYPES},
      "SCR_GlazeInk": {t: data[(t,p)]["stock_glaze_t"]    / data[(t,p)]["cons_glaze_tgg"]    for t in TYPES},
    }
def PI_of(p):
    return {
      "PsI_Energy":  {t: round(data[(t,p)]["firstchoice_rate"]/(data[(t,p)]["energia_processo_MJ_m2"]/1000.0), 3) for t in TYPES},
      "PsI_Yield":   {t: data[(t,p)]["resa_m2_m2"] for t in TYPES},
      "PsI_Through": {t: float(data[(t,p)]["throughput_m2_h"]) for t in TYPES},
    }
def OCR_of(p):
    return {
      "OCR_Flex":  {t: round(data[(t,p)]["flex_QP_Nmm2"]/data[(t,p)]["flex_AT_Nmm2"], 3) for t in TYPES},
      "OCR_Break": {t: round(data[(t,p)]["break_QP_N"]  /data[(t,p)]["break_AT_N"],   3) for t in TYPES},
      "OCR_Surf":  {t: round(data[(t,p)]["surf_QP_pct"] /data[(t,p)]["surf_AT_pct"],  3) for t in TYPES},
    }
SCR, PI, OCR = SCR_of("t"), PI_of("t"), OCR_of("t")
DIM   = {"IOA": SCR, "OP": PI, "TQ": OCR}
DIM_1 = {"IOA": SCR_of("t-1"), "OP": PI_of("t-1"), "TQ": OCR_of("t-1")}

# ============================================================================
# 4. PRIMARIA: z-score tra le 3 tipologie, pesi uguali
# ============================================================================
def zscore_dim(dd):
    z = {}
    for ind, vals in dd.items():
        a = np.array([vals[t] for t in TYPES], float); m, s = a.mean(), a.std(ddof=0)
        z[ind] = {t: (vals[t]-m)/s if s > 0 else 0.0 for t in TYPES}
    return z
def subindex_z(dd):
    z = zscore_dim(dd); inds = list(dd); w = 1.0/len(inds)
    return {t: sum(w*z[i][t] for i in inds) for t in TYPES}, z
IOAI_z, zIOA = subindex_z(SCR); OPI_z, zOP = subindex_z(PI); TQI_z, zTQ = subindex_z(OCR)
PTSI_z = {t: (IOAI_z[t]+OPI_z[t]+TQI_z[t])/3.0 for t in TYPES}

# ============================================================================
# 5. SECONDARIA: scoring 1-5 + AHP
# ============================================================================
TH = {  # 4 confini di classe -> punteggio 1..5 (higher = better)
 "SCR_RawMat": [20,30,40,55], "SCR_Finished": [12,18,25,35], "SCR_GlazeInk": [12,18,25,32],
 "PsI_Energy": [8.0,12.0,16.0,20.0], "PsI_Yield": [0.93,0.945,0.955,0.965], "PsI_Through": [200,350,500,620],
 "OCR_Flex": [1.15,1.30,1.40,1.50], "OCR_Break": [1.2,1.6,2.2,3.5], "OCR_Surf": [1.005,1.015,1.025,1.03],
}
def score15(ind, x):
    th = TH[ind]
    return 1 if x < th[0] else 2 if x < th[1] else 3 if x < th[2] else 4 if x < th[3] else 5
def ahp(M):
    M = np.array(M, float); n = M.shape[0]; gm = np.prod(M, 1)**(1.0/n); w = gm/gm.sum()
    lam = (M @ w / w).mean(); CI = (lam-n)/(n-1); RI = {2:0,3:0.58,4:0.90}[n]; return w, lam, CI, (CI/RI if RI else 0)
AHP_IOA=[[1,1,2],[1,1,2],[.5,.5,1]]; AHP_OP=[[1,2,2],[.5,1,1],[.5,1,1]]
AHP_TQ=[[1,1,3],[1,1,3],[1/3,1/3,1]]; AHP_DIM=[[1,.5,1/3],[2,1,.5],[3,2,1]]
wIOA,_,_,crI=ahp(AHP_IOA); wOP,_,_,crO=ahp(AHP_OP); wTQ,_,_,crT=ahp(AHP_TQ); wDIM,lamD,ciD,crD=ahp(AHP_DIM)
def sdim(dd, wv):
    inds=list(dd); return {t: sum(wv[i]*score15(inds[i], dd[inds[i]][t]) for i in range(len(inds))) for t in TYPES}
def ptsi5(dims):
    a=sdim(dims["IOA"],wIOA); b=sdim(dims["OP"],wOP); c=sdim(dims["TQ"],wTQ)
    return a,b,c,{t: wDIM[0]*a[t]+wDIM[1]*b[t]+wDIM[2]*c[t] for t in TYPES}
S_IOA,S_OP,S_TQ,PTSI_5 = ptsi5(DIM)
_,_,_,PTSI_5_1 = ptsi5(DIM_1)
TII = {t: round((PTSI_5[t]/PTSI_5_1[t]-1)*100, 2) for t in TYPES}
def ptsi5_equal(dims):
    def sd(d):
        inds=list(d); w=1.0/len(inds); return {t: sum(w*score15(i,d[i][t]) for i in inds) for t in TYPES}
    a,b,c=sd(dims["IOA"]),sd(dims["OP"]),sd(dims["TQ"]); return {t:(a[t]+b[t]+c[t])/3 for t in TYPES}
PTSI_5_eq = ptsi5_equal(DIM)

# ============================================================================
# 6. SENSIBILITA
# ============================================================================
def rankstr(x): return " > ".join(sorted(x, key=lambda t: x[t], reverse=True))
SENS = {
 "z-score + pesi uguali (PRIMARIO)": PTSI_z,
 "z-score + pesi AHP (TQ>OP>IOA)": {t: wDIM[0]*IOAI_z[t]+wDIM[1]*OPI_z[t]+wDIM[2]*TQI_z[t] for t in TYPES},
 "scoring 1-5 + AHP": PTSI_5,
 "scoring 1-5 + pesi uguali": PTSI_5_eq,
 "z-score focus OP (0.25/0.50/0.25)": {t: .25*IOAI_z[t]+.5*OPI_z[t]+.25*TQI_z[t] for t in TYPES},
 "z-score focus disponib./qualita (0.40/0.20/0.40)": {t: .4*IOAI_z[t]+.2*OPI_z[t]+.4*TQI_z[t] for t in TYPES},
}

RES = {"SCR":SCR,"PI":PI,"OCR":OCR,"IOAI_z":IOAI_z,"OPI_z":OPI_z,"TQI_z":TQI_z,"PTSI_z":PTSI_z,
       "S_IOA":S_IOA,"S_OP":S_OP,"S_TQ":S_TQ,"PTSI_5":PTSI_5,"PTSI_5_t1":PTSI_5_1,"PTSI_5_eq":PTSI_5_eq,"TII":TII,
       "AHP":{"wIOA":[round(x,4) for x in wIOA],"wOP":[round(x,4) for x in wOP],"wTQ":[round(x,4) for x in wTQ],
              "wDIM":[round(x,4) for x in wDIM],"CR":{"IOA":round(crI,4),"OP":round(crO,4),"TQ":round(crT,4),"DIM":round(crD,4)}},
       "SENS":{k:{"valori":{t:round(v[t],3) for t in TYPES},"ranking":rankstr(v)} for k,v in SENS.items()},
       "e_spec":{t:data[(t,"t")]["energia_processo_MJ_m2"] for t in TYPES},
       "massa":{t:data[(t,"t")]["massa_kg_m2"] for t in TYPES}}
with open(os.path.join(OUT, "RP7.4_results.json"), "w") as f: json.dump(RES, f, indent=1, default=float)

# ============================================================================
# 7. OUTPUT: weights.xlsx e calculation_log.xlsx
# ============================================================================
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "AHP_within_dim"
def dump(ws, name, M, w, cr, inds):
    ws.append([name]); ws.append([""]+inds)
    for i, r in enumerate(M): ws.append([inds[i]]+list(r))
    ws.append(["peso_AHP"]+[round(x,4) for x in w]); ws.append(["CR", round(cr,4), "OK (<=0.10)" if cr<=0.10 else "RIVEDERE"]); ws.append([])
dump(ws,"IOA (SCR)",AHP_IOA,wIOA,crI,["SCR_RawMat","SCR_Finished","SCR_GlazeInk"])
dump(ws,"OP (PsI)",AHP_OP,wOP,crO,["PsI_Energy","PsI_Yield","PsI_Through"])
dump(ws,"TQ (OCR)",AHP_TQ,wTQ,crT,["OCR_Flex","OCR_Break","OCR_Surf"])
_autosize(ws)
ws=wb.create_sheet("AHP_between_dim"); ws.append(["","IOA","OP","TQ"])
for i,r in enumerate(AHP_DIM): ws.append([["IOA","OP","TQ"][i]]+list(r))
ws.append(["peso_AHP"]+[round(x,4) for x in wDIM]); ws.append(["lambda_max",round(lamD,4)])
ws.append(["CI",round(ciD,4)]); ws.append(["RI(n=3)",0.58]); ws.append(["CR",round(crD,4),"OK (<=0.10)" if crD<=0.10 else "RIVEDERE"]); _autosize(ws)
ws=wb.create_sheet("Soglie_scoring_1_5"); ws.append(["indicatore","classe1_<","classe2_<","classe3_<","classe4_<","classe5_>=","direzione"])
for ind,th in TH.items(): ws.append([ind,th[0],th[1],th[2],th[3],th[3],"higher=better"])
_style_header(ws); _autosize(ws)
ws=wb.create_sheet("NOTE"); ws.append(["Pesi/soglie provvisori, da confermare in workshop produzione/qualita/manutenzione."])
wb.save(os.path.join(OUT,"RP7.4_weights.xlsx"))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "calculation_log"
ws.append(["result_id","report_table","tipologia","periodo","variabile","formula","input_source","output","unita","note","versione"])
rid=[0]
def add(tbl,t,per,var,formula,src,out,unit):
    rid[0]+=1; ws.append([f"P{rid[0]:03d}",tbl,t,per,var,formula,src,round(out,4) if isinstance(out,float) else out,unit,"provvisorio - in corso di consolidamento","beta-1.0"])
for ind in SCR:
    for t in TYPES: add("T2",t,"t",ind,"AS/AC","dataset:INPUT",SCR[ind][t],"giorni")
for ind in PI:
    for t in TYPES: add("T3",t,"t",ind,"ROU/RIN","dataset:INPUT",PI[ind][t],"varie")
for ind in OCR:
    for t in TYPES: add("T4",t,"t",ind,"QP/AT","dataset:INPUT",OCR[ind][t],"adim")
for t in TYPES: add("T5",t,"t","IOAI_z","mean(w*z(SCR))","zscore",IOAI_z[t],"adim")
for t in TYPES: add("T5",t,"t","OPI_z","mean(w*z(PsI))","zscore",OPI_z[t],"adim")
for t in TYPES: add("T5",t,"t","TQI_z","mean(w*z(OCR))","zscore",TQI_z[t],"adim")
for t in TYPES: add("T6",t,"t","P-TSI_z","mean(IOAI,OPI,TQI)","T5",PTSI_z[t],"adim (z, primario)")
for t in TYPES: add("T7",t,"t","S_IOA","sum(w_AHP*score15)","weights",S_IOA[t],"[1-5]")
for t in TYPES: add("T7",t,"t","S_OP","sum(w_AHP*score15)","weights",S_OP[t],"[1-5]")
for t in TYPES: add("T7",t,"t","S_TQ","sum(w_AHP*score15)","weights",S_TQ[t],"[1-5]")
for t in TYPES: add("T8",t,"t","P-TSI_5","sum(wDIM*S_dim)","T7",PTSI_5[t],"[1-5] (secondario)")
for t in TYPES: add("T8",t,"t-1","P-TSI_5","sum(wDIM*S_dim)","T7",PTSI_5_1[t],"[1-5]")
for t in TYPES: add("T9",t,"t-1,t","TII","(P-TSI_5_t/P-TSI_5_t1-1)*100","T8",TII[t],"%")
_style_header(ws); _autosize(ws)
ws=wb.create_sheet("NOTE"); ws.append(["Serie 2023-2024 provvisorie, in corso di consolidamento, ancorate a EPD e serie RP6.x/RP7.3."])
ws.append(["A fine progetto: sostituire i dati nel dataset e rieseguire; struttura di calcolo invariata."])
wb.save(os.path.join(OUT,"RP7.4_calculation_log.xlsx"))

# ============================================================================
# 8. FIGURE
# ============================================================================
try:
    import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
    lab={"T1_7.4mm":"T1 · 7,4 mm","T2_8.2mm":"T2 · 8,2 mm","T3_20mm":"T3 · 20 mm"}
    col={"T1_7.4mm":"#4C72B0","T2_8.2mm":"#DD8452","T3_20mm":"#55A868"}
    plt.rcParams.update({"font.size":11})
    x=np.arange(3); w=0.26
    fig,ax=plt.subplots(figsize=(8,4.6))
    for i,t in enumerate(TYPES):
        ax.bar(x+(i-1)*w,[IOAI_z[t],OPI_z[t],TQI_z[t]],w,label=lab[t],color=col[t],edgecolor="white",lw=0.6)
    ax.axhline(0,color="#444",lw=0.8); ax.set_xticks(x)
    ax.set_xticklabels(["IOA\n(disponibilità)","OP\n(performance)","TQ\n(qualità)"])
    ax.set_ylabel("Sotto-indice normalizzato (z-score)"); ax.set_title("P-TSA — profilo dimensionale per tipologia (z-score)")
    ax.legend(frameon=False,ncol=3,loc="upper center",bbox_to_anchor=(0.5,-0.12)); ax.spines[["top","right"]].set_visible(False)
    fig.tight_layout(); fig.savefig(os.path.join(OUT,"RP7.4_fig1_profilo_dimensionale.png"),dpi=150,bbox_inches="tight")
    fig,ax=plt.subplots(figsize=(7,4.6))
    vals=[PTSI_5[t] for t in TYPES]; tii=[TII[t] for t in TYPES]
    bars=ax.bar([lab[t] for t in TYPES],vals,color=[col[t] for t in TYPES],edgecolor="white",width=0.6)
    for b,v,ti in zip(bars,vals,tii):
        ax.text(b.get_x()+b.get_width()/2,v+0.05,f"{v:.2f}\n(TII +{ti:.1f}%)",ha="center",va="bottom",fontsize=10)
    ax.set_ylim(0,5); ax.set_ylabel("P-TSI  [scala 1–5]"); ax.set_title("P-TSI per tipologia (scoring 1–5 + AHP) e miglioramento annuo (TII)")
    ax.spines[["top","right"]].set_visible(False); ax.axhline(3,color="#bbb",ls="--",lw=0.8)
    ax.text(2.4,3.03,"soglia medio-alta",fontsize=8,color="#888")
    fig.tight_layout(); fig.savefig(os.path.join(OUT,"RP7.4_fig2_ptsi_tii.png"),dpi=150,bbox_inches="tight")
except Exception as e:
    print("[fig] skip:", e)

print("rank primario:", rankstr(PTSI_z))
print("P-TSI_z:", {t:round(PTSI_z[t],3) for t in TYPES})
print("P-TSI_5:", {t:round(PTSI_5[t],3) for t in TYPES}, "TII:", TII)
print("AHP CR:", RES["AHP"]["CR"])
print("OK -> dataset, weights, calculation_log, figures")
