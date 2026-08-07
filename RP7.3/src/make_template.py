# -*- coding: utf-8 -*-
"""Template Excel di raccolta dati RP7.3 (precompilato 2022 rif. + 2023-2025).
   Allineato agli input del motore: Energia + termini dei 4 moduli (MJ) + coefficienti + AHP."""
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src import core
from src.ahp import TRIAL_MATRIX, DIMS
BASE=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT=os.path.join(BASE,"output","RP7.3_data_collection_2023-2025.xlsx")
HF=PatternFill("solid",fgColor="0B5A3C"); HFONT=Font(color="FFFFFF",bold=True)
REFF=PatternFill("solid",fgColor="EDEDED")
THIN=Border(*[Side(style="thin",color="BFBFBF")]*4)
ALLY=[core.REF_YEAR]+core.YEARS
def _h(ws,row=1):
    for c in ws[row]:
        c.fill=HF; c.font=HFONT; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def _grid(ws):
    for r in ws.iter_rows():
        for c in r: c.border=THIN
def sheet(wb,title,headers,rows,widths):
    ws=wb.create_sheet(title); ws.append(headers)
    for row in rows: ws.append(row)
    _h(ws); _grid(ws)
    for i,wd in enumerate(widths): ws.column_dimensions[chr(65+i)].width=wd
    # evidenzia righe anno riferimento
    for r in range(2,ws.max_row+1):
        if str(ws.cell(r,2).value)==str(core.REF_YEAR):
            for c in ws[r]: c.fill=REFF
    return ws
def build():
    wb=openpyxl.Workbook(); ist=wb.active; ist.title="Istruzioni"
    msg=[("RP7.3 — Scheda di raccolta dati (serie 2023–2025)",True,13),
         ("",False,11),
         ("Anno 2022 = riferimento (baseline) per le differenze dei moduli; 2023–2025 = serie di lavoro.",False,11),
         ("I valori inseriti sono provvisori e in corso di consolidamento con le serie storiche definitive.",False,11),
         ("",False,11),
         ("Come compilare:",True,11),
         ("1) 'Unita': anagrafica e produzione annua (m²).",False,11),
         ("2) 'Energia': consumi primari — gas naturale (Nm³) ed energia elettrica (kWh), inclusa autoproduzione.",False,11),
         ("3) 'Moduli_*': termini exergetici di ciascun modulo in MJ (vedi manuali -J).",False,11),
         ("4) 'Coefficienti' e 'AHP': librerie di conversione e matrice dei pesi.",False,11),
         ("",False,11),
         ("Rigenerazione: sostituire i valori e rieseguire  python3 -m src.run_all  (stessa struttura di calcolo).",False,11),
         ("Convenzione unità: calcolo interno in MJ; output in GJ (÷1.000). Exergia gas = chimica (b_fuel).",False,11)]
    for i,(t,b,s) in enumerate(msg,1):
        c=ist.cell(i,1,t); c.font=Font(bold=b,size=s)
    ist.column_dimensions["A"].width=110
    # Unita
    sheet(wb,"Unita",["plant","descrizione","produzione_m2"],
          [[p,core.PLANTS[p]["desc"],core.PLANTS[p]["P"]] for p in core.PLANTS],[12,32,16])
    # Energia
    er=[]
    for p in core.PLANTS:
        for y in ALLY:
            e=core.energy_split(p,y); er.append([p,y,round(e["V_gas_Nm3"],0),round(e["kWh"],0)])
    sheet(wb,"Energia",["plant","year","V_gas_Nm3","E_el_kWh"],er,[12,8,16,16])
    # Moduli
    def terms(cols): 
        rr=[]
        for p in core.PLANTS:
            for y in ALLY:
                tv=core.term_value(p,y); rr.append([p,y]+[round(tv[c],0) for c in cols])
        return rr
    sheet(wb,"Moduli_TEI",["plant","year","loss_MTS_MJ","loss_MTO_MJ","inv_MJ","qual_MTS_MJ","qual_MTO_MJ"],
          terms(["loss_MTS","loss_MTO","inv","qual_MTS","qual_MTO"]),[12,8,14,14,12,13,13])
    sheet(wb,"Moduli_EFA",["plant","year","RI_MJ","IEQ_MJ","WEX_MJ","CIRC_MJ"],
          terms(["RI","IEQ","WEX","CIRC"]),[12,8,16,14,12,12])
    sheet(wb,"Moduli_EcoFA",["plant","year","VA_MJ","econ_in_MJ","INV_MJ"],
          terms(["VA","econ_in","INV"]),[12,8,16,14,14])
    sheet(wb,"Moduli_SFA",["plant","year","SV_MJ","train_MJ","lost_MJ","CO2_MJ"],
          terms(["SV","train","lost","CO2"]),[12,8,14,12,12,12])
    # Coefficienti
    sheet(wb,"Coefficienti",["code","description","value","unit","source","year","boundary","method","confidence","version"],
          [["EL_EX","Exergia elettrica",core.KEL,"MJ/kWh","fisica",2025,"—","fisico","A","1.0"],
           ["GAS_EX","Exergia chimica gas",core.B_FUEL,"MJ/Nm3","exergia chimica","2025","CTG","termod.","B","1.0"],
           ["IMP_CO2","CO2-eq -> J",441868.0,"MJ/tCO2e","modello","2024","—","modello","B","1.0"],
           ["ECO_VA","VA EUR->MJ",3.2,"MJ/EUR","settoriale","2024","—","top-down","B","1.0"],
           ["ECO_IN","Input econ. EUR->MJ",4.5,"MJ/EUR","fornitori","2024","—","bottom-up","B","1.0"],
           ["LAB_H","Exergia oraria lavoro",3.5,"MJ/h","modello","2024","—","modello","B","1.0"]],
          [10,26,10,10,14,8,10,12,11,9])
    # AHP
    wsa=wb.create_sheet("AHP"); wsa.append([""]+DIMS)
    for i,d in enumerate(DIMS): wsa.append([d]+[round(x,4) for x in TRIAL_MATRIX[i]])
    _h(wsa); _grid(wsa)
    for c in wsa["A"]: c.font=Font(bold=True)
    for col in "ABCDE": wsa.column_dimensions[col].width=14
    wb.save(OUT); return OUT
if __name__=="__main__":
    print("template:",build())
