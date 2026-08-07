# -*- coding: utf-8 -*-
"""Genera gli input DIMOSTRATIVI (CSV) e le librerie controllate (coefficients_master.xlsx, ahp_weights.xlsx).
   TUTTI i valori sono DIMOSTRATIVI / NON VALIDATI (confidence C). Da sostituire con dati primari ERP/MES + E2C."""
import os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from src import core
from src.ahp import weights_and_consistency, TRIAL_MATRIX, DIMS

BASE=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INP=os.path.join(BASE,"input"); COE=os.path.join(BASE,"coefficients")
HDR_FILL=PatternFill("solid",fgColor="0B5A3C"); HDR_FONT=Font(color="FFFFFF",bold=True)

def _style_header(ws,row=1):
    for c in ws[row]:
        c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=Alignment(horizontal="center",vertical="center")

def write_inputs():
    # energy_exergy.csv
    with open(os.path.join(INP,"energy_exergy.csv"),"w",newline="") as f:
        w=csv.writer(f); w.writerow(["plant","scenario","production_m2","V_gas_Nm3","E_el_kWh","Ex_fuel_MJ","Ex_el_MJ","Ex_ref_MJ","Ex_useful_MJ","confidence","note"])
        for p in core.PLANTS:
            for s in core.SCENARIOS:
                e=core.energy_split(p,s)
                w.writerow([p,s,core.PLANTS[p]["P"],round(e["V_gas_Nm3"],1),round(e["kWh"],1),
                            round(e["Ex_fuel_MJ"],1),round(e["Ex_el_MJ"],1),round(e["Ex_ref_MJ"],1),
                            round(e["Ex_useful_MJ"],1),"C","DIMOSTRATIVO"])
    # module_terms.csv
    with open(os.path.join(INP,"module_terms.csv"),"w",newline="") as f:
        w=csv.writer(f); w.writerow(["plant","scenario","module","term","value_MJ","kind","confidence","note"])
        modmap={"loss_MTS":"TEI-J","loss_MTO":"TEI-J","inv":"TEI-J","qual_MTS":"TEI-J","qual_MTO":"TEI-J",
                "RI":"EFA-J","IEQ":"EFA-J","WEX":"EFA-J","CIRC":"EFA-J",
                "VA":"EcoFA-J","econ_in":"EcoFA-J","INV":"EcoFA-J",
                "SV":"SFA-J","train":"SFA-J","lost":"SFA-J","CO2":"SFA-J"}
        for p in core.PLANTS:
            for s in core.SCENARIOS:
                tv=core.term_value(p,s)
                for t,v in tv.items():
                    w.writerow([p,s,modmap[t],t,round(v,1),core.TERM_KIND[t],"C","DIMOSTRATIVO"])

def write_coefficients():
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Coefficients"
    ws.append(["code","description","value","unit","source","year","boundary","method","confidence","version"])
    rows=[
        ["EL_EX","Exergia elettrica (fattore unitario)",core.KEL,"MJ/kWh","fisica (1 kWh=3.6 MJ)",2025,"—","fisico","A","1.0"],
        ["GAS_EX","Exergia chimica metano",core.B_FUEL,"MJ/Nm3","proxy ~1.04*LHV","2025","CTG","proxy","C","1.0-DEMO"],
        ["MAT_CLAY","Argilla (CED cradle-to-gate)",1.5,"MJ/kg","EPD fornitore [placeholder]","2024","CTG","LCA","C","1.0-DEMO"],
        ["MAT_FELD","Feldspato",2.1,"MJ/kg","DB LCA [placeholder]","2024","CTG","LCA","C","1.0-DEMO"],
        ["IMP_CO2","CO2-eq -> Joule",441868.0,"MJ/tCO2e","SFA+ alpha","2024","—","modello","B","1.0"],
        ["ECO_VA","Valore aggiunto EUR->MJ",3.2,"MJ/EUR","intensita macro [placeholder]","2024","—","top-down","C","1.0-DEMO"],
        ["ECO_IN","Input economici EUR->MJ",4.5,"MJ/EUR","dati fornitori [placeholder]","2024","—","bottom-up","C","1.0-DEMO"],
        ["LAB_H","Exergia oraria lavoro",3.5,"MJ/h","EEA+ alpha (metab.+cogn.)","2024","—","proxy","C","1.0-DEMO"],
    ]
    for r in rows: ws.append(r)
    _style_header(ws)
    for col,wd in zip("ABCDEFGHIJ",[10,34,12,10,26,8,10,12,11,10]): ws.column_dimensions[col].width=wd
    note=wb.create_sheet("NOTE")
    note["A1"]="ATTENZIONE: valori DIMOSTRATIVI / NON VALIDATI (prevalente confidenza C)."
    note["A2"]="Da sostituire con coefficienti primari/EPD approvati dal progetto prima di ogni uso decisionale."
    note["A1"].font=Font(bold=True,color="C00000")
    wb.save(os.path.join(COE,"coefficients_master.xlsx"))

def write_ahp():
    w,lam,CI,CR=weights_and_consistency(TRIAL_MATRIX)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Pairwise_Matrix"
    ws.append([""]+DIMS)
    for i,d in enumerate(DIMS):
        ws.append([d]+[round(x,4) for x in TRIAL_MATRIX[i]])
    _style_header(ws)
    for c in ws["A"]: c.font=Font(bold=True)
    ws2=wb.create_sheet("Weights_Consistency")
    ws2.append(["dimension","weight_AHP"]);
    for d,val in zip(DIMS,w): ws2.append([d,round(float(val),4)])
    ws2.append(["SUM",round(float(sum(w)),4)])
    ws2.append([]); ws2.append(["lambda_max",round(lam,4)]); ws2.append(["CI",round(CI,4)])
    ws2.append(["RI(n=4)",0.90]); ws2.append(["CR",round(CR,4)])
    ws2.append(["esito","CONSISTENTE (CR<=0.10)" if CR<=0.10 else "REVISIONE (CR>0.10)"])
    _style_header(ws2)
    ws3=wb.create_sheet("NOTE")
    ws3["A1"]="Matrice di confronto a coppie DI PROVA (DIMOSTRATIVA)."; ws3["A1"].font=Font(bold=True,color="C00000")
    ws3["A2"]="Da sostituire con i giudizi del panel tecnico-progettuale (scala Saaty 1-9)."
    for col,wd in zip("ABCDE",[16,12,12,12,12]):
        ws.column_dimensions[col].width=wd; ws2.column_dimensions[col].width=wd
    wb.save(os.path.join(COE,"ahp_weights.xlsx"))

if __name__=="__main__":
    write_inputs(); write_coefficients(); write_ahp()
    print("Input CSV + coefficients_master.xlsx + ahp_weights.xlsx generati (DIMOSTRATIVI).")
