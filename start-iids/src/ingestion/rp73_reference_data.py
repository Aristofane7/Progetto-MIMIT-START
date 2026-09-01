"""Loader for the real (provisional) RP7.3 reference data workbook.

Spec ref: ADR-012. Reads `data/reference/RP7.3_data_collection_20232025.xlsx`
(`Unita`, `Energia`, `Moduli_TEI/EFA/EcoFA/SFA` sheets) into plain, per-plant/year
dataclasses. This module does NOT touch `Coefficienti`/`AHP` (those are already
transcribed into `config/coefficients/rp73_provisional_2026.yaml` and
`config/weights/eea_ahp_rp73.yaml` — load those via
`src.core.coefficients.load_coefficient_set` / `src.core.weights.load_weight_set`).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

BASELINE_YEAR = 2022  # fixed reference year for every plant (ADR-012)


@dataclass(frozen=True)
class PlantMaster:
    plant_id: str
    description: str
    production_m2: float


@dataclass(frozen=True)
class EnergyRow:
    plant_id: str
    year: int
    v_gas_nm3: float
    e_el_kwh: float


@dataclass(frozen=True)
class TEIModuleRow:
    plant_id: str
    year: int
    loss_mts_mj: float
    loss_mto_mj: float
    inv_mj: float
    qual_mts_mj: float
    qual_mto_mj: float


@dataclass(frozen=True)
class EFAModuleRow:
    plant_id: str
    year: int
    ri_mj: float
    ieq_mj: float
    wex_mj: float
    circ_mj: float


@dataclass(frozen=True)
class EcoFAModuleRow:
    plant_id: str
    year: int
    va_mj: float
    econ_in_mj: float
    inv_mj: float


@dataclass(frozen=True)
class SFAModuleRow:
    plant_id: str
    year: int
    sv_mj: float
    train_mj: float
    lost_mj: float
    co2_mj: float


@dataclass(frozen=True)
class RP73ReferenceData:
    plants: dict[str, PlantMaster]
    energy: dict[tuple[str, int], EnergyRow]
    tei: dict[tuple[str, int], TEIModuleRow]
    efa: dict[tuple[str, int], EFAModuleRow]
    ecofa: dict[tuple[str, int], EcoFAModuleRow]
    sfa: dict[tuple[str, int], SFAModuleRow]

    def plant_years(self) -> list[tuple[str, int]]:
        """All (plant_id, year) pairs excluding the baseline year itself."""
        return sorted(k for k in self.energy if k[1] != BASELINE_YEAR)


def _rows(ws) -> list[tuple]:
    values = list(ws.iter_rows(values_only=True))
    return values[1:]  # skip header row


def load_rp73_reference_data(path: str | Path) -> RP73ReferenceData:
    wb = openpyxl.load_workbook(path, data_only=True)

    plants = {
        row[0]: PlantMaster(plant_id=row[0], description=row[1], production_m2=row[2])
        for row in _rows(wb["Unita"])
    }
    energy = {
        (row[0], row[1]): EnergyRow(plant_id=row[0], year=row[1], v_gas_nm3=row[2], e_el_kwh=row[3])
        for row in _rows(wb["Energia"])
    }
    tei = {
        (row[0], row[1]): TEIModuleRow(
            plant_id=row[0], year=row[1], loss_mts_mj=row[2], loss_mto_mj=row[3],
            inv_mj=row[4], qual_mts_mj=row[5], qual_mto_mj=row[6],
        )
        for row in _rows(wb["Moduli_TEI"])
    }
    efa = {
        (row[0], row[1]): EFAModuleRow(
            plant_id=row[0], year=row[1], ri_mj=row[2], ieq_mj=row[3], wex_mj=row[4], circ_mj=row[5],
        )
        for row in _rows(wb["Moduli_EFA"])
    }
    ecofa = {
        (row[0], row[1]): EcoFAModuleRow(
            plant_id=row[0], year=row[1], va_mj=row[2], econ_in_mj=row[3], inv_mj=row[4],
        )
        for row in _rows(wb["Moduli_EcoFA"])
    }
    sfa = {
        (row[0], row[1]): SFAModuleRow(
            plant_id=row[0], year=row[1], sv_mj=row[2], train_mj=row[3], lost_mj=row[4], co2_mj=row[5],
        )
        for row in _rows(wb["Moduli_SFA"])
    }

    return RP73ReferenceData(plants=plants, energy=energy, tei=tei, efa=efa, ecofa=ecofa, sfa=sfa)


def load_rp73_calculation_log(path: str | Path) -> dict[tuple[str, int, str], float]:
    """Load `RP7.3_calculation_log.xlsx` (`calculation_log` sheet) into
    ``{(plant_id, year, variable): output}`` — the golden regression targets
    used by `tests/regression/test_rp73_calculation_log.py`."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["calculation_log"]
    log: dict[tuple[str, int, str], float] = {}
    for row in _rows(ws):
        _result_id, _report_table, plant, year, variable, *_rest, output, _unit, _date, _version = row
        log[(plant, year, variable)] = output
    return log
